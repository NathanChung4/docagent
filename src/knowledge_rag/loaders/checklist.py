"""Excel checklist loader.

Treats each row of the checklist as its own Document so individual items can be
retrieved and cited directly. The first sheet is read; the first row is assumed
to be column headers.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from knowledge_rag.loaders.base import Loader
from knowledge_rag.models import Document, SourceType


class ChecklistLoader(Loader):
    """Loads a single .xlsx checklist file into one Document per row."""

    def load(self, path: Path) -> list[Document]:
        if not path.exists() or not path.is_file():
            return []
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h) if h is not None else "" for h in header_row]

        # Pick the first column as the natural item identifier (e.g., "component").
        id_col = headers[0] if headers else "item"

        docs: list[Document] = []
        for row_index, row in enumerate(rows_iter, start=2):
            # strict=False: trailing empty cells in real spreadsheets can make rows shorter than headers.
            row_dict: dict[str, Any] = {h: v for h, v in zip(headers, row, strict=False) if h}
            if not any(v is not None and v != "" for v in row_dict.values()):
                continue  # skip blank rows
            item_id = str(row_dict.get(id_col, f"row_{row_index}"))

            # Render row as readable text for embedding.
            text = "\n".join(f"{k}: {v}" for k, v in row_dict.items() if v is not None)

            docs.append(
                Document(
                    source_type=SourceType.CHECKLIST,
                    title=f"{item_id} (checklist)",
                    content=text,
                    uri=f"{path.resolve()}#row={row_index}",
                    metadata={
                        "filename": path.name,
                        "row_index": row_index,
                        "item_id": item_id,
                        "fields": row_dict,
                    },
                )
            )
        return docs
