"""Generic helpers shared by domain-pack tools.

Domain isolation: this module never names a specific domain. Domain packs hold
the spec dicts (KNOWN_SPECS, ENUM_SPECS, PRIMARY_METRIC) and pass them in as
arguments. Anything domain-specific belongs in `domains/<name>/tools/`, not here.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from knowledge_rag.tools.base import ToolValidationError


@dataclass(frozen=True)
class ParamSpec:
    """Numeric parameter constraint: type + inclusive range + default."""

    type: type  # int or float
    lo: float
    hi: float
    default: float


@dataclass(frozen=True)
class EnumSpec:
    """String-enum parameter constraint."""

    choices: frozenset[str]
    default: str | None = None


# A spec table maps parameter name -> constraint (numeric or enum).
SpecTable = dict[str, ParamSpec | EnumSpec]


# --- validation --------------------------------------------------------------


def validate_params(item_name: str, params: dict[str, Any], spec: SpecTable) -> None:
    """Raise ToolValidationError on missing / unknown / out-of-range / wrong-type params."""
    if not params:
        raise ToolValidationError(
            f"No parameters supplied for '{item_name}'. Expected one of: {', '.join(sorted(spec))}."
        )
    for key, value in params.items():
        constraint = spec.get(key)
        if constraint is None:
            raise ToolValidationError(
                f"Unknown parameter '{key}' for '{item_name}'. Allowed: {', '.join(sorted(spec))}."
            )
        if isinstance(constraint, EnumSpec):
            if value not in constraint.choices:
                raise ToolValidationError(
                    f"Parameter '{key}' must be one of {sorted(constraint.choices)}, got {value!r}."
                )
            continue
        # bool is a subclass of int — reject so True doesn't pass as 1.
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ToolValidationError(
                f"Parameter '{key}' must be {constraint.type.__name__}, got {type(value).__name__}."
            )
        if constraint.type is int and not isinstance(value, int):
            raise ToolValidationError(f"Parameter '{key}' must be int, got {type(value).__name__}.")
        if not (constraint.lo <= value <= constraint.hi):
            raise ToolValidationError(
                f"Parameter '{key}'={value} is outside allowed range "
                f"[{constraint.lo}, {constraint.hi}] for '{item_name}'."
            )


# --- CSV write ---------------------------------------------------------------


def write_kv_csv(
    output_dir: Path,
    stem: str,
    params: dict[str, Any],
    *,
    key_header: str = "param",
    value_header: str = "value",
) -> Path:
    """Write a sorted key/value CSV; return the absolute path written."""
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(UTC).strftime("%Y%m%dT%H%M%S")
    out_path = output_dir / f"{stem}_{ts}.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([key_header, value_header])
        for key in sorted(params):
            writer.writerow([key, params[key]])
    return out_path.resolve()


# --- sweep-report helpers ----------------------------------------------------


def _coerce_number(value: Any) -> float | str:
    """CSVs are all strings; surface numerics for downstream sorting and stats."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def find_latest_report(directory: Path, item_name: str) -> Path:
    """Locate the latest sweep CSV for `item_name`. Raise ToolValidationError if none."""
    if not directory.exists():
        raise ToolValidationError(f"Sweep reports directory {directory} does not exist.")
    candidates = sorted(directory.glob(f"{item_name}*sweep*.csv"))
    if not candidates:
        candidates = sorted(directory.glob(f"*{item_name}*.csv"))
    if not candidates:
        raise ToolValidationError(f"No sweep report found for '{item_name}' in {directory}.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def read_csv_rows(csv_path: Path) -> list[dict[str, Any]]:
    """Read a CSV into list-of-dicts with numeric coercion."""
    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = [{k: _coerce_number(v) for k, v in raw.items()} for raw in reader]
    return rows


_PICKERS = {"max": max, "min": min}


def pick_winning_run(
    rows: list[dict[str, Any]],
    metric: str,
    direction: str,
) -> dict[str, Any]:
    """Return the row whose `metric` is best (max or min)."""
    picker = _PICKERS.get(direction)
    if picker is None:
        raise ToolValidationError(f"direction must be 'max' or 'min', got {direction!r}.")
    numeric_rows = [(r, r[metric]) for r in rows if isinstance(r.get(metric), (int, float))]
    if not numeric_rows:
        raise ToolValidationError(f"No rows have a numeric '{metric}' value to optimize.")
    return picker(numeric_rows, key=lambda rv: rv[1])[0]


def compute_param_ranges(
    rows: list[dict[str, Any]],
    *,
    exclude: set[str],
) -> dict[str, dict[str, float]]:
    """Per numeric column not in `exclude`, return {min, max} across all rows."""
    if not rows:
        return {}
    ranges: dict[str, dict[str, float]] = {}
    for col in rows[0]:
        if col in exclude:
            continue
        values = [r[col] for r in rows if isinstance(r.get(col), (int, float))]
        if values:
            ranges[col] = {"min": min(values), "max": max(values)}
    return ranges


# --- xlsx lookup -------------------------------------------------------------


def lookup_xlsx_row(xlsx_path: Path, key: str) -> dict[str, Any] | None:
    """Return the row whose first column == key (case-insensitive), or None.

    Values are stringified so the result is JSON-serializable straight to the
    agent (datetime cells in particular don't survive json.dumps otherwise).
    """
    if not xlsx_path.exists():
        raise ToolValidationError(f"Checklist file not found at {xlsx_path}.")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        raise ToolValidationError(f"Checklist {xlsx_path.name} has no active sheet.")

    rows = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows)
    except StopIteration:
        return None
    headers = [str(h) if h is not None else "" for h in headers_raw]
    if not headers:
        return None
    id_col = headers[0]

    target = key.strip().lower()
    for row in rows:
        # strict=False: trailing empty cells in real spreadsheets can make rows shorter than headers.
        row_dict = {h: v for h, v in zip(headers, row, strict=False) if h}
        if str(row_dict.get(id_col, "")).strip().lower() == target:
            return {k: (str(v) if v is not None else None) for k, v in row_dict.items()}
    return None
